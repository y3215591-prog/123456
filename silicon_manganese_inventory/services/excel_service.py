import openpyxl
import re
from silicon_manganese_inventory.dao.database import DatabaseManager
from silicon_manganese_inventory.dao.base_dao import (
    CustomerDAO, SalesOrderDAO, DailyShipmentDAO,
    SupplierDAO, SpecDAO, LocationDAO,
    OperationLogDAO,
)
from silicon_manganese_inventory.services.seal_service import (
    SealService, SealInsufficientError, SealStatusError,
)
from silicon_manganese_inventory.dao.outbound_dao import OutboundDAO


class ExcelService:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def import_sales_orders(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        customer_dao = CustomerDAO(self.db)
        order_dao = SalesOrderDAO(self.db)
        from silicon_manganese_inventory.dao.base_dao import SpecDAO
        spec_dao = SpecDAO(self.db)

        stats = {"new_customers": 0, "updated_customers": 0,
                 "new_specs": 0, "imported_orders": 0, "skipped": 0}
        order_no_idx = self._find_col(headers, ["销售订单号", "order_no"])
        customer_code_idx = self._find_col(headers, ["客户", "客户代码", "customer_code"])
        customer_name_idx = self._find_col(headers, ["客户名称", "customer_name"])
        material_desc_idx = self._find_col(headers, ["物料描述", "material_desc"])

        if order_no_idx is None:
            raise ValueError("Excel 缺少销售订单号列")

        for row in ws.iter_rows(min_row=2, values_only=True):
            order_no = str(row[order_no_idx] or "").strip()
            if not order_no:
                stats["skipped"] += 1
                continue

            customer_code = str(row[customer_code_idx] or "").strip() if customer_code_idx is not None else ""
            customer_name = str(row[customer_name_idx] or "").strip() if customer_name_idx is not None else ""
            material_desc = str(row[material_desc_idx] or "").strip() if material_desc_idx is not None else ""

            if customer_code and customer_name:
                existing = customer_dao.get_by_code(customer_code)
                if existing:
                    if existing["name"] != customer_name:
                        customer_dao.update(existing["id"], name=customer_name)
                        stats["updated_customers"] += 1
                else:
                    customer_dao.create(code=customer_code, name=customer_name)
                    stats["new_customers"] += 1

            if material_desc and spec_dao:
                existing_spec = spec_dao.get_by_name(material_desc.split(",")[0].strip())
                if not existing_spec:
                    spec_dao.create(material_desc.split(",")[0].strip())
                    stats["new_specs"] += 1

            kwargs = {}
            for col_name, col_idx_map in [
                ("order_no", order_no_idx),
                ("line_no", self._find_col(headers, ["销售订单行号", "line_no"])),
                ("customer_code", customer_code_idx),
                ("customer_name", customer_name_idx),
                ("contract_ref", self._find_col(headers, ["合同参考", "contract_ref"])),
                ("contract_no", self._find_col(headers, ["销售合同号", "contract_no"])),
                ("material_code", self._find_col(headers, ["物料编码", "material_code"])),
                ("material_desc", material_desc_idx),
                ("delivery_start", self._find_col(headers, ["交货开始日期", "delivery_start"])),
                ("delivery_end", self._find_col(headers, ["交货截止日期", "delivery_end"])),
                ("delivery_address", self._find_col(headers, ["送货地址", "delivery_address"])),
                ("quantity", self._find_col(headers, ["数量", "quantity"])),
                ("unit", self._find_col(headers, ["单位", "unit"])),
                ("factory_code", self._find_col(headers, ["工厂", "工厂代码", "factory_code"])),
                ("factory_name", self._find_col(headers, ["工厂名称", "factory_name"])),
                ("pickup_method", self._find_col(headers, ["提货方式", "pickup_method"])),
            ]:
                idx = col_idx_map
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        kwargs[col_name] = val

            existing_order = order_dao.get_by_order_no(order_no)
            if not existing_order:
                order_dao.create(**kwargs)
                stats["imported_orders"] += 1

        wb.close()
        return stats

    def import_daily_shipments(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        dao = DailyShipmentDAO(self.db)

        headers = []
        for cell in ws[1]:
            val = str(cell.value or "").strip()
            headers.append(val)

        stats = {"imported": 0, "skipped": 0}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                stats["skipped"] += 1
                continue
            kwargs = {}
            col_map = {
                "seq_no": 0, "shipment_date": 1, "plate_no": 2,
                "customer_code": 3, "customer_name": 4,
                "sales_order_no": 5, "material_name": 6, "spec": 7,
                "batch_no": 8, "load_quantity": 9,
                "gross_weight": 10, "tare_weight": 11, "net_weight": 12,
                "customer_received_weight": 13, "remark": 14,
            }
            for key, idx in col_map.items():
                if idx < len(row) and row[idx] is not None:
                    kwargs[key] = row[idx]
            if kwargs.get("seq_no") is not None:
                dao.create(**kwargs)
                stats["imported"] += 1

        wb.close()
        return stats

    def _find_col(self, headers, names):
        for i, h in enumerate(headers):
            if h in names:
                return i
        return None

    def import_sales_orders_from_rows(self, rows):
        customer_dao = CustomerDAO(self.db)
        order_dao = SalesOrderDAO(self.db)
        from silicon_manganese_inventory.dao.base_dao import SpecDAO
        spec_dao = SpecDAO(self.db)

        if not rows:
            return {"new_customers": 0, "updated_customers": 0,
                    "new_specs": 0, "imported_orders": 0, "skipped": 0}

        headers = [str(c or "").strip() for c in rows[0]]
        stats = {"new_customers": 0, "updated_customers": 0,
                 "new_specs": 0, "imported_orders": 0, "skipped": 0}
        order_no_idx = self._find_col(headers, ["销售订单号", "order_no"])
        customer_code_idx = self._find_col(headers, ["客户", "客户代码", "customer_code"])
        customer_name_idx = self._find_col(headers, ["客户名称", "customer_name"])
        material_desc_idx = self._find_col(headers, ["物料描述", "material_desc"])

        if order_no_idx is None:
            raise ValueError("Excel 缺少销售订单号列")

        for row in rows[1:]:
            order_no = str(row[order_no_idx] or "").strip()
            if not order_no:
                stats["skipped"] += 1
                continue
            customer_code = str(row[customer_code_idx] or "").strip() if customer_code_idx is not None else ""
            customer_name = str(row[customer_name_idx] or "").strip() if customer_name_idx is not None else ""
            material_desc = str(row[material_desc_idx] or "").strip() if material_desc_idx is not None else ""

            if customer_code and customer_name:
                existing = customer_dao.get_by_code(customer_code)
                if existing:
                    if existing["name"] != customer_name:
                        customer_dao.update(existing["id"], name=customer_name)
                        stats["updated_customers"] += 1
                else:
                    customer_dao.create(code=customer_code, name=customer_name)
                    stats["new_customers"] += 1

            if material_desc and spec_dao:
                existing_spec = spec_dao.get_by_name(material_desc.split(",")[0].strip())
                if not existing_spec:
                    spec_dao.create(material_desc.split(",")[0].strip())
                    stats["new_specs"] += 1

            kwargs = {}
            for col_name, col_idx_map in [
                ("order_no", order_no_idx),
                ("line_no", self._find_col(headers, ["销售订单行号", "line_no"])),
                ("customer_code", customer_code_idx),
                ("customer_name", customer_name_idx),
                ("contract_ref", self._find_col(headers, ["合同参考", "contract_ref"])),
                ("contract_no", self._find_col(headers, ["销售合同号", "contract_no"])),
                ("material_code", self._find_col(headers, ["物料编码", "material_code"])),
                ("material_desc", material_desc_idx),
                ("delivery_start", self._find_col(headers, ["交货开始日期", "delivery_start"])),
                ("delivery_end", self._find_col(headers, ["交货截止日期", "delivery_end"])),
                ("delivery_address", self._find_col(headers, ["送货地址", "delivery_address"])),
                ("quantity", self._find_col(headers, ["数量", "quantity"])),
                ("unit", self._find_col(headers, ["单位", "unit"])),
                ("factory_code", self._find_col(headers, ["工厂", "工厂代码", "factory_code"])),
                ("factory_name", self._find_col(headers, ["工厂名称", "factory_name"])),
                ("pickup_method", self._find_col(headers, ["提货方式", "pickup_method"])),
            ]:
                idx = col_idx_map
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        kwargs[col_name] = val

            existing_order = order_dao.get_by_order_no(order_no)
            if not existing_order:
                order_dao.create(**kwargs)
                stats["imported_orders"] += 1
        return stats

    def import_daily_shipments_from_rows(self, rows):
        dao = DailyShipmentDAO(self.db)
        if not rows:
            return {"imported": 0, "skipped": 0}
        stats = {"imported": 0, "skipped": 0}
        col_map = {
            "seq_no": 0, "shipment_date": 1, "plate_no": 2,
            "customer_code": 3, "customer_name": 4,
            "sales_order_no": 5, "material_name": 6, "spec": 7,
            "batch_no": 8, "load_quantity": 9,
            "gross_weight": 10, "tare_weight": 11, "net_weight": 12,
            "customer_received_weight": 13, "remark": 14,
        }
        for row in rows[1:]:
            if row[0] is None:
                stats["skipped"] += 1
                continue
            kwargs = {}
            for key, idx in col_map.items():
                if idx < len(row) and row[idx] is not None:
                    kwargs[key] = row[idx]
            if kwargs.get("seq_no") is not None:
                dao.create(**kwargs)
                stats["imported"] += 1
        return stats

    def import_mes_shipments(self, file_path, operator="系统导入", auto_deduct=True):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        shipment_dao = DailyShipmentDAO(self.db)
        customer_dao = CustomerDAO(self.db)
        spec_dao = SpecDAO(self.db)
        location_dao = LocationDAO(self.db)
        outbound_dao = OutboundDAO(self.db)
        seal_service = SealService(self.db)
        log_dao = OperationLogDAO(self.db)

        max_seq = shipment_dao.get_max_seq_no()
        stats = {"imported": 0, "skipped": 0, "duplicate": 0,
                 "outbound_created": 0, "seal_shipped": 0}
        imported_orders = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 24:
                stats["skipped"] += 1
                continue

            car_no = str(row[5] or "").strip()
            outbound_no = str(row[1] or "").strip()

            if not car_no or outbound_no in imported_orders:
                if outbound_no in imported_orders:
                    stats["duplicate"] += 1
                else:
                    stats["skipped"] += 1
                continue
            imported_orders.add(outbound_no)

            sales_order = str(row[2] or "").strip()
            customer_name = str(row[4] or "").strip()

            driver = str(row[7] or "").strip()
            driver_phone = str(row[8] or "").strip()

            try:
                gross_weight = float(row[11] or 0)
            except (ValueError, TypeError):
                gross_weight = 0
            try:
                tare_weight = float(row[12] or 0)
            except (ValueError, TypeError):
                tare_weight = 0
            try:
                net_weight = float(row[13] or 0)
            except (ValueError, TypeError):
                net_weight = 0

            outbound_time = str(row[18] or "").strip()
            shipment_date = outbound_time[:10] if outbound_time else ""

            material_detail = str(row[23] or "").strip()

            batches, batch_entries, spec_name, material_name, total_to = \
                self._parse_material_detail(material_detail)

            max_seq += 1
            remark_parts = []
            if driver:
                remark_parts.append(f"司机:{driver}")
            if driver_phone:
                remark_parts.append(f"电话:{driver_phone}")
            locations = [e[2] for e in batch_entries if e[2]]
            if locations:
                remark_parts.append(f"库位:{','.join(locations)}")
            remark = "; ".join(remark_parts)

            ship_id = None
            try:
                ship_id = shipment_dao.create(
                    seq_no=max_seq,
                    shipment_date=shipment_date,
                    plate_no=car_no,
                    customer_name=customer_name,
                    sales_order_no=sales_order,
                    material_name=material_name,
                    spec=spec_name,
                    batch_no=",".join(batches) if batches else "",
                    load_quantity=total_to,
                    gross_weight=gross_weight,
                    tare_weight=tare_weight,
                    net_weight=net_weight,
                    remark=remark,
                )
                stats["imported"] += 1

                log_dao.log("import_shipment", "daily_shipments", ship_id,
                            f"导入发货: 车号={car_no}, 订单={sales_order}, "
                            f"客户={customer_name}, 数量={total_to}TO, "
                            f"批次={','.join(batches) if batches else ''}",
                            operator)

                if customer_name:
                    existing = customer_dao.get_by_name(customer_name)
                    if not existing:
                        cust_id = customer_dao.create(name=customer_name)
                        log_dao.log("auto_create_customer", "customers", cust_id,
                                    f"自动创建客户: {customer_name}", operator)

                if auto_deduct and batch_entries and spec_name:
                    self._auto_deduct_inventory(
                        ship_id, shipment_date, customer_name, sales_order,
                        car_no, spec_name, material_name,
                        batch_entries, customer_dao, spec_dao,
                        location_dao, outbound_dao, seal_service,
                        log_dao, stats, operator)

            except Exception:
                stats["skipped"] += 1

        wb.close()

        log_dao.log("import_mes_complete", "daily_shipments", None,
                    f"MES导入完成: 导入{stats['imported']}条, "
                    f"跳过{stats['skipped']}条, 重复{stats['duplicate']}条, "
                    f"出库单{stats['outbound_created']}个, "
                    f"发货铅封{stats['seal_shipped']}个",
                    operator)
        return stats

    def _auto_deduct_inventory(self, ship_id, shipment_date, customer_name,
                               sales_order, plate_no, spec_name, material_name,
                               batch_entries, customer_dao, spec_dao,
                               location_dao, outbound_dao, seal_service,
                               log_dao, stats, operator):
        cust = customer_dao.get_by_name(customer_name)
        if not cust:
            return
        customer_id = cust["id"]

        spec = spec_dao.get_by_name(spec_name)
        if not spec:
            spec_id = spec_dao.create(name=spec_name)
            spec = spec_dao.get(spec_id)
        spec_id = spec["id"]

        total_qty = sum(qty for _, qty, _ in batch_entries)
        outbound_id, order_no = outbound_dao.create_outbound(
            date=shipment_date, customer_id=customer_id,
            sales_order_no=sales_order, spec_id=spec_id,
            quantity=total_qty, plate_no=plate_no,
            operator=operator, remark=f"自动扣减-{material_name}",
        )
        stats["outbound_created"] += 1
        log_dao.log("create_outbound", "outbound_orders", outbound_id,
                    f"自动出库: {order_no}, 客户={customer_name}, "
                    f"数量={total_qty}TO, 车号={plate_no}",
                    operator)

        all_seal_codes = []
        all_batches = []
        for batch_no, qty, location in batch_entries:
            if qty <= 0:
                continue
            try:
                seal_start, seal_end, seal_codes = seal_service.ship_seals(
                    qty, batch_no=batch_no, location_code=location,
                    outbound_id=outbound_id)
                all_seal_codes.extend(seal_codes)
                all_batches.append(batch_no)
                stats["seal_shipped"] += qty
                log_dao.log("ship_seals", "seal_numbers", outbound_id,
                            f"发货铅封: {seal_start}~{seal_end} ({qty}个), "
                            f"批次={batch_no}, 库位={location}, 订单={order_no}",
                            operator)
            except (SealInsufficientError, SealStatusError) as e:
                log_dao.log("ship_failed", "seal_numbers", outbound_id,
                            f"发货失败: {e}, 批次={batch_no}, 库位={location}, "
                            f"需求{qty}个",
                            operator)

        with self.db.get_connection() as conn:
            conn.execute(
                "UPDATE outbound_orders SET seal_start=?, seal_end=?, batch_nos=? WHERE id=?",
                (all_seal_codes[0] if all_seal_codes else "",
                 all_seal_codes[-1] if all_seal_codes else "",
                 ",".join(all_batches) if all_batches else "",
                 outbound_id),
            )
            conn.execute(
                "UPDATE daily_shipments SET outbound_id=?, seal_codes=? WHERE id=?",
                (outbound_id,
                 ",".join(all_seal_codes) if all_seal_codes else "",
                 ship_id),
            )

    def _parse_material_detail(self, detail):
        batches = []
        batch_entries = []
        spec_name = ""
        material_name = ""
        total_to = 0

        for line in detail.split("\n"):
            seg = line.strip()
            if not seg:
                continue

            seg_clean = re.sub(r'\s*合格\s*', ' ', seg)
            seg_clean = re.sub(r'\s*粒径:\S+\s*', ' ', seg_clean)
            seg_clean = re.sub(r'\s+', ' ', seg_clean).strip()

            if not spec_name:
                parts = seg_clean.split(",")
                if len(parts) >= 1:
                    spec_name = parts[0].strip()
                if len(parts) >= 2:
                    material_name = parts[1].strip()

            batch_match = re.search(r'批次:(\d{10})', seg_clean)
            batch_num = batch_match.group(1) if batch_match else ""

            to_match = re.search(r'(\d{1,3})TO', seg_clean)
            to_qty = int(to_match.group(1)) if to_match else 0

            loc_match = re.search(r'\s+([A-Z]\d{1,2})\s*$', seg_clean)
            location = loc_match.group(1) if loc_match else ""

            total_to += to_qty
            if batch_num:
                batches.append(batch_num)
                batch_entries.append((batch_num, to_qty, location))

        return batches, batch_entries, spec_name, material_name, total_to

    def import_customers_from_sales_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        customer_dao = CustomerDAO(self.db)
        stats = {"imported": 0, "skipped": 0}

        header_row = [str(c.value or "").strip() for c in ws[1]]
        code_idx = self._find_col(header_row, ["客户", "customer_code"])
        name_idx = self._find_col(header_row, ["客户名称", "customer_name"])
        if code_idx is None or name_idx is None:
            wb.close()
            raise ValueError("Excel 缺少客户或客户名称列")

        seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            code = str(row[code_idx] or "").strip() if code_idx < len(row) else ""
            name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
            if not code or not name:
                stats["skipped"] += 1
                continue
            if code in seen:
                stats["skipped"] += 1
                continue
            seen.add(code)
            try:
                existing = customer_dao.get_by_code(code)
                if not existing:
                    customer_dao.create(code=code, name=name)
                    stats["imported"] += 1
                else:
                    stats["skipped"] += 1
            except Exception:
                stats["skipped"] += 1

        wb.close()
        return stats

    def import_inventory_balance(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        results = []
        total_balance = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            if not row or len(row) < 9:
                continue
            location = str(row[2] or "").strip()
            batch_no = str(row[3] or "").strip()
            balance = row[8]
            if not batch_no:
                continue
            try:
                balance = float(balance) if balance else 0
            except (ValueError, TypeError):
                balance = 0
            if balance <= 0:
                continue
            total_balance += balance
            results.append({
                "location": location,
                "batch_no": batch_no,
                "balance": balance,
                "material_code": row[0],
                "material_name": row[1],
                "unit": row[4],
                "row": row_idx,
            })
        wb.close()
        return {
            "items": results,
            "total_balance": total_balance,
            "count": len(results),
        }


class ExportService:
    def __init__(self, db: DatabaseManager):
        self.db = db

    def export_inventory(self, file_path):
        from silicon_manganese_inventory.services.report_service import ReportService
        service = ReportService(self.db)
        rows = service.get_inventory_report()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成品库存"
        ws.append(["批次号", "库位", "结存(吨)", "最近入库日期", "化验结果", "铅封号明细"])
        for row in rows:
            ws.append([
                row["batch_no"], row["location_code"], row["balance"],
                row["last_inbound_date"], row["overall_result"] or "",
                row["seal_list"] or "",
            ])
        wb.save(file_path)

    def export_daily_shipments(self, file_path, include_seals=True, **filters):
        dao = DailyShipmentDAO(self.db)
        rows = dao.list(**filters)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "每日发货明细"
        headers = ["序号", "发货日期", "车牌", "客户代码", "客户名称",
                   "销售订单号", "批次号", "装车吨数",
                   "毛重", "皮重", "净重", "客户收货净重", "备注"]
        col_map = ["seq_no", "shipment_date", "plate_no",
                   "customer_code", "customer_name",
                   "sales_order_no", "batch_no", "load_quantity",
                   "gross_weight", "tare_weight", "net_weight",
                   "customer_received_weight", "remark"]
        if include_seals:
            headers.insert(11, "铅封号")
            col_map.insert(11, "seal_codes")
        ws.append(headers)
        for row in rows:
            ws.append([row[k] for k in col_map])
        wb.save(file_path)

    def export_order_summary(self, file_path):
        from silicon_manganese_inventory.services.report_service import ReportService
        service = ReportService(self.db)
        rows = service.get_order_summary()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "订单装车汇总"
        ws.append(["订单号", "客户代码", "客户名称", "物料名称", "规格", "单位",
                    "订单量", "交货截止日期", "已发量", "待发量", "完成率",
                    "提货方式", "发货进度"])
        for row in rows:
            ws.append([
                row["order_no"], row["customer_code"], row["customer_name"],
                row["material_name"], row["spec"], row["unit"],
                row["order_quantity"], row["delivery_end"],
                row["shipped_quantity"], row["pending_quantity"],
                row["completion_rate"], row["pickup_method"], row["warning"],
            ])
        wb.save(file_path)

    def export_pre_inbound(self, file_path, **filters):
        from silicon_manganese_inventory.dao.inbound_dao import InboundDAO
        dao = InboundDAO(self.db)
        rows = dao.list_pre_inbound(**filters)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "预入库管理"
        ws.append(["预入库单号", "日期", "批次号", "数量(吨)", "库位",
                    "铅封号段", "化验状态", "入库状态", "操作人", "备注"])
        for row in rows:
            seal_range = ""
            if row["seal_start"]:
                seal_range = f"{row['seal_start']}~{row['seal_end']}"
            lab_status = "已化验" if row["lab_status"] == "tested" else "待化验"
            inbound_status = "已入库" if row.get("inbound_status") == "confirmed" else "未入库"
            ws.append([
                row["order_no"], row["date"], row["batch_no"], row["quantity"],
                row["location_code"], seal_range, lab_status, inbound_status,
                row["operator"], row["remark"] or "",
            ])
        wb.save(file_path)

    def export_seals(self, file_path, batch_id=None, status=None):
        from silicon_manganese_inventory.dao.seal_dao import SealDAO
        dao = SealDAO(self.db)
        rows = dao.list_all(status=status)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "铅封号管理"
        ws.append(["铅封号", "号段", "状态", "批次号", "库位",
                    "预入库单", "入库单", "出库单", "更新时间"])
        for row in rows:
            status_map = {"unused": "未使用", "pre_allocated": "预分配",
                          "in_stock": "在库", "shipped": "已发货"}
            ws.append([
                row["seal_code"],
                row.get("batch_code", ""),
                status_map.get(row["status"], row["status"]),
                row.get("batch_no", ""),
                row.get("location_code", ""),
                row.get("pre_inbound_order", "") or "",
                row.get("inbound_order", "") or "",
                row.get("outbound_order", "") or "",
                row.get("updated_at", ""),
            ])
        wb.save(file_path)

    def export_outbound(self, file_path, **filters):
        from silicon_manganese_inventory.dao.outbound_dao import OutboundDAO
        dao = OutboundDAO(self.db)
        rows = dao.list_outbound(**filters)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出库发货"
        ws.append(["出库单号", "日期", "批次号", "品名规格", "数量(吨)",
                    "客户", "销售订单号", "合同号", "铅封号范围", "车牌号",
                    "操作人", "备注"])
        for row in rows:
            seal_range = ""
            if row["seal_start"]:
                seal_range = f"{row['seal_start']}~{row['seal_end']}"
            ws.append([
                row["order_no"], row["date"], row.get("batch_nos", ""),
                row.get("spec_name", "") or "", row["quantity"],
                row.get("customer_name", "") or "",
                row["sales_order_no"] or "", row.get("contract_no", "") or "",
                seal_range, row.get("plate_no", "") or "",
                row["operator"], row["remark"] or "",
            ])
        wb.save(file_path)
